"""Safety/privacy guardrails for hosted Streamlit use."""

from __future__ import annotations

import io
import zipfile

from openpyxl import load_workbook
import pandas as pd
import pytest

import streamlit_app as app


class OversizeUpload:
    """Upload-like object whose content must not be read."""

    name = "large.csv"
    size = app.TABLE_FILE_HARD_LIMIT_BYTES

    def getvalue(self):  # pragma: no cover - failure path
        raise AssertionError("oversize upload was read")


class SizedUpload:
    def __init__(self, name: str, size: int):
        self.name = name
        self.size = size


def test_read_flexible_table_refuses_declared_oversize_upload_before_reading():
    with pytest.raises(ValueError, match="not parsed"):
        app.read_flexible_table("", OversizeUpload(), header=True)


def test_prepare_download_frames_public_mode_excludes_identifier_heavy_tables():
    frames = {
        "summary": pd.DataFrame({"Metric": ["N"], "Value": [10]}),
        "scorefile": pd.DataFrame({"Person": ["S01"], "Observed": [4]}),
        "measures": pd.DataFrame({
            "Facet": ["Person", "Rater"],
            "Level": ["S01", "R1"],
            "Estimate": [0.2, -0.1],
        }),
        "bias_table": pd.DataFrame({
            "Facet1": ["Person"],
            "Level1": ["S01"],
            "Facet2": ["Rater"],
            "Level2": ["R1"],
        }),
    }

    public_frames = app.prepare_download_frames_for_privacy(frames, public_export_mode=True)

    assert "summary" in public_frames
    assert "scorefile" not in public_frames
    assert "bias_table" not in public_frames
    assert public_frames["measures"]["Facet"].tolist() == ["Rater"]
    manifest = public_frames["export_privacy_manifest"]
    assert "scorefile" in manifest["Frame"].tolist()
    assert "excluded_public_mode" in manifest["Status"].tolist()


def test_prepare_download_frames_private_mode_keeps_complete_tables():
    frames = {
        "scorefile": pd.DataFrame({"Person": ["S01"], "Observed": [4]}),
    }

    private_frames = app.prepare_download_frames_for_privacy(frames, public_export_mode=False)

    assert "scorefile" in private_frames
    assert private_frames["scorefile"]["Person"].tolist() == ["S01"]
    assert "export_privacy_manifest" in private_frames


def test_export_zip_names_are_sanitized_and_sheet_names_unique():
    frames = {
        "../unsafe/path": pd.DataFrame({"x": [1]}),
        "a" * 40: pd.DataFrame({"x": [2]}),
        "a" * 39 + "b": pd.DataFrame({"x": [3]}),
    }

    zip_bytes = app.cached_tables_zip(frames, app.frames_fingerprint(frames))
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
    assert all(".." not in name and "/" not in name and "\\" not in name for name in names)
    assert any(name.endswith(".csv") for name in names)

    workbook = load_workbook(io.BytesIO(app.to_excel_bytes(frames)), read_only=True)
    assert len(workbook.sheetnames) == 3
    assert len(set(workbook.sheetnames)) == 3
    assert all(len(name) <= 31 for name in workbook.sheetnames)


def test_estimation_resource_preflight_blocks_large_hosted_runs():
    df = pd.DataFrame({
        "Person": [f"P{i}" for i in range(app.ESTIMATION_HOSTED_MAX_OBS + 1)],
        "Rater": ["R1"] * (app.ESTIMATION_HOSTED_MAX_OBS + 1),
        "Task": ["T1"] * (app.ESTIMATION_HOSTED_MAX_OBS + 1),
        "Score": [1] * (app.ESTIMATION_HOSTED_MAX_OBS + 1),
    })

    preflight = app.build_estimation_resource_preflight(
        data=df,
        person_col="Person",
        score_col="Score",
        facet_cols=["Rater", "Task"],
        model_type="RSM",
        est_method="JMLE",
        maxit=400,
        quad_points=15,
        bias_mode="Skip",
        bias_pairs_available=[],
        selected_bias_pair=None,
        compute_strict_marginal=False,
        strict_marginal_pairwise=False,
        strict_marginal_max_pair_cells=400,
        generate_figure_exports=False,
    )

    assert preflight["block"] is True
    assert app.estimation_resource_preflight_status(preflight) == "Block"
    assert app.should_render_estimation_resource_preflight(preflight) is True
    assert any(row["Metric"] == "Observations" and row["Status"] == "Block" for row in preflight["rows"])


def test_estimation_resource_preflight_suppresses_ok_noise_but_surfaces_review():
    small = pd.DataFrame({
        "Person": ["P1", "P2", "P3", "P4"],
        "Rater": ["R1", "R1", "R2", "R2"],
        "Task": ["T1", "T2", "T1", "T2"],
        "Score": [1, 2, 3, 4],
    })
    ok_preflight = app.build_estimation_resource_preflight(
        data=small,
        person_col="Person",
        score_col="Score",
        facet_cols=["Rater", "Task"],
        model_type="RSM",
        est_method="JMLE",
        maxit=400,
        quad_points=15,
        bias_mode="Skip",
        bias_pairs_available=[],
        selected_bias_pair=None,
        compute_strict_marginal=False,
        strict_marginal_pairwise=False,
        strict_marginal_max_pair_cells=400,
        generate_figure_exports=False,
    )
    assert app.estimation_resource_preflight_status(ok_preflight) == "OK"
    assert app.should_render_estimation_resource_preflight(ok_preflight) is False

    large_preview = pd.DataFrame({
        "Person": [f"P{i % 500}" for i in range(25_001)],
        "Rater": ["R1"] * 25_001,
        "Task": ["T1"] * 25_001,
        "Score": [1] * 25_001,
    })
    review_preflight = app.build_estimation_resource_preflight(
        data=large_preview,
        person_col="Person",
        score_col="Score",
        facet_cols=["Rater", "Task"],
        model_type="RSM",
        est_method="JMLE",
        maxit=3_001,
        quad_points=15,
        bias_mode="Skip",
        bias_pairs_available=[],
        selected_bias_pair=None,
        compute_strict_marginal=False,
        strict_marginal_pairwise=False,
        strict_marginal_max_pair_cells=400,
        generate_figure_exports=False,
    )
    assert app.estimation_resource_preflight_status(review_preflight) == "Review"
    assert app.should_render_estimation_resource_preflight(review_preflight) is True


def test_posterior_upload_and_payload_preflights_block_oversize_inputs():
    upload_errors, _ = app.posterior_upload_preflight([
        SizedUpload("too_large.nc", app.POSTERIOR_MAX_FILE_BYTES + 1)
    ])
    assert upload_errors

    payload_errors, _ = app.posterior_payload_preflight({
        "n_chains": 4,
        "n_draws": app.POSTERIOR_MAX_DRAWS_PER_CHAIN + 1,
        "parameter_names": ["theta"] * (app.POSTERIOR_MAX_PARAMETERS + 1),
    })
    assert payload_errors
